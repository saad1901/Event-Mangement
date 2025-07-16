
from django.shortcuts import render
from django.http import HttpResponse
from app.models import Participant, Tournament
import csv
from django.contrib.auth.decorators import login_required
from django.utils import timezone
import io


def downloads_home(request):
    context = {
        'active_tab': "admin_downloads",
        'tournaments': Tournament.objects.all(),
        'status_choices': Participant.STATUS_CHOICES,
    }
    return render(request, 'admin/downloads.html',context)

@login_required
def export_participants_csv(request):
    # Accept POSTed filters
    tournament_id = request.POST.get('tournament_id', 'all')
    status = request.POST.get('status', 'all')
    download_type = request.POST.get('download_type', 'csv')

    participants = Participant.objects.select_related('tournament').all()
    if tournament_id and tournament_id != 'all':
        participants = participants.filter(tournament_id=tournament_id)
    if status and status != 'all':
        participants = participants.filter(status=status)

    # Determine if any participant is chess and build headers accordingly
    chess_headers = [
        'FIDE ID', 'National ID', 'Title', 'FIDE Rating', 'National Rating',
        # 'Rapid Rating', 'Blitz Rating',
        'Section', 'Federation', 'Club/Academy',
        # 'Is Arbiter',
        # 'Previous Tournaments', 'Achievements'
    ]
    base_headers = [
        'Full Name','Phone', 'WhatsApp','Status', 'Tournament', 'Registration Date'
    ]
    # Check if any participant is chess
    has_chess = any(
        p.tournament and hasattr(p.tournament, 'category') and getattr(p.tournament.category, 'name', '').lower() == 'chess' for p in participants
    )
    headers = base_headers + (chess_headers if has_chess else [])
    rows = []
    for p in participants:
        row = [
            p.full_name,
            p.phone,
            p.wp,
            p.status,
            p.tournament.title if p.tournament else '',
            p.registration_date.strftime('%Y-%m-%d %H:%M:%S') if p.registration_date else '',
        ]
        is_chess = p.tournament and hasattr(p.tournament, 'category') and getattr(p.tournament.category, 'name', '').lower() == 'chess'
        if has_chess:
            if is_chess and hasattr(p, 'chess_profile') and p.chess_profile:
                chess = p.chess_profile
                row += [
                    chess.fide_id or '',
                    chess.national_id or '',
                    chess.title or '',
                    chess.fide_rating or '',
                    chess.national_rating or '',
                    # chess.rapid_rating or '',
                    # chess.blitz_rating or '',
                    chess.section or '',
                    chess.federation or '',
                    chess.club_academy or '',
                    # 'Yes' if chess.is_arbiter else 'No',
                    # chess.previous_tournaments or '',
                    # chess.achievements or '',
                ]
            else:
                row += [''] * len(chess_headers)
        rows.append(row)

    # Get tournament info for header
    tournament_name = 'All Tournaments'
    tournament_date = ''
    if tournament_id and tournament_id != 'all':
        t = Tournament.objects.filter(id=tournament_id).first()
        if t:
            tournament_name = t.title
            tournament_date = str(t.start_date)
    # Prepare file name
    safe_tournament_name = tournament_name.replace(' ', '') if tournament_name != 'All Tournaments' else 'AllTournaments'
    safe_status = status.replace(' ', '') if status and status != 'all' else ''
    file_name_base = safe_tournament_name
    if safe_status:
        file_name_base += f'_{safe_status}'

    if download_type == 'excel':
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl is required for Excel export.', status=500)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        ws['A1'] = f'Tournament: {tournament_name}'
        ws['A2'] = f'Date: {tournament_date}'
        for row in rows:
            ws.append(row)
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name_base}.xlsx"'
        return response
    elif download_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Flowable
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
        except ImportError:
            return HttpResponse('reportlab is required for PDF export.', status=500)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph(f'Tournament: {tournament_name}', styles['Title']))
        if tournament_date:
            elements.append(Paragraph(f'Date: {tournament_date}', styles['Normal']))
        elements.append(Spacer(1, 18))

        # Prepare participant blocks
        participant_blocks = []
        for row in rows:
            lines = []
            for i, value in enumerate(row):
                label = headers[i]
                # Color the Status field
                if label.lower() == 'status':
                    status_str = str(value).strip().lower()
                    if status_str == 'confirmed':
                        color = 'green'
                    elif status_str == 'registered':
                        color = 'blue'
                    elif status_str == 'rejected':
                        color = 'red'
                    else:
                        color = 'black'
                    lines.append(f"<b>{label}:</b> <font color='{color}'><b>{value if value else '-'}</b></font>")
                else:
                    lines.append(f"<b>{label}:</b> {value if value else '-'}")
            participant_blocks.append(Paragraph('<br/>'.join(lines), styles['Normal']))

        # Arrange blocks in pairs (2 per row)
        table_data = []
        i = 0
        while i < len(participant_blocks):
            left = participant_blocks[i]
            right = participant_blocks[i+1] if i+1 < len(participant_blocks) else ''
            table_data.append([left, right])
            i += 2

        # Table style: add padding, vertical alignment, etc.
        table_style = TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ])

        # Add table and black lines after each row
        for idx, row in enumerate(table_data):
            t = Table([row], colWidths=[250, 250])
            t.setStyle(table_style)
            elements.append(t)
            elements.append(Spacer(1, 8))
            # Add a solid black line after each row
            class BlackLine(Flowable):
                def __init__(self, width=510, thickness=2):
                    super().__init__()
                    self.width = width
                    self.thickness = thickness
                def draw(self):
                    self.canv.setStrokeColor(colors.black)
                    self.canv.setLineWidth(self.thickness)
                    self.canv.line(0, 0, self.width, 0)
            elements.append(BlackLine())
            elements.append(Spacer(1, 8))

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{file_name_base}.pdf"'
        return response
    else:  # CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name_base}.csv"'
        writer = csv.writer(response)
        writer.writerow([f'Tournament: {tournament_name}'])
        writer.writerow([f'Date: {tournament_date}'])
        writer.writerow([])
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return response